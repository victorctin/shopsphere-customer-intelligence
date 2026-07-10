"""Build the M7 Excel executive workbook from the gold layer, then validate:
the script re-opens the written file and asserts its headline cells against
freshly queried SQL values. Exit 1 on any mismatch. Read-only against MySQL.

KPI definitions are copied verbatim from the M3 gate (python/03_gold/run_gold.py)
and the RFM / A/B notebooks so the workbook reports the contract numbers.
"""
import sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from scipy import stats
from sqlalchemy import bindparam, text

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from config import settings
from config.db import get_engine

OUT_PATH = settings.PROJECT_ROOT / "reports" / "shopsphere_executive_workbook.xlsx"
REL_TOLERANCE = 0.005          # same 0.5% band as the M3 gate
AOV_EXPECTED = 93.96           # M3 gate expectation
AB_LIFT_EXPECTED = 0.151       # docs/calibration_report.md, A/B relative lift

TITLE_FONT = Font(bold=True, size=14)
HEADER_FONT = Font(bold=True)


def fetch_scalar(conn, sql: str) -> float:
    return float(conn.execute(text(sql)).scalar_one())


def compute_kpis(engine) -> dict:
    """Headline KPIs using the exact M3 gate SQL (run_gold.py)."""
    with engine.connect() as conn:
        kpis = {
            "total_revenue": fetch_scalar(
                conn, "SELECT SUM(order_revenue) FROM gold_order_revenue "
                      "WHERE order_status = 'completed'"),
            "completed_orders": fetch_scalar(
                conn, "SELECT COUNT(*) FROM gold_order_revenue "
                      "WHERE order_status = 'completed'"),
            "aov": fetch_scalar(
                conn, "SELECT AVG(order_revenue) FROM gold_order_revenue "
                      "WHERE order_status = 'completed'"),
            "buyers": fetch_scalar(
                conn, "SELECT COUNT(*) FROM gold_customer_pareto"),
            "one_time_rate": fetch_scalar(
                conn, "SELECT AVG(orders_all = 1) FROM gold_customer_summary "
                      "WHERE orders_all > 0"),
            "abandonment": fetch_scalar(
                conn, "SELECT 1 - MAX(CASE WHEN event_type = 'purchase' "
                      "THEN events END) / MAX(CASE WHEN event_type = "
                      "'add_to_cart' THEN events END) FROM gold_funnel"),
            "top20_share": fetch_scalar(
                conn, "SELECT SUM(CASE WHEN revenue_rank <= "
                      "FLOOR(buyer_count * 0.2) THEN revenue_completed "
                      "ELSE 0 END) / SUM(revenue_completed) "
                      "FROM gold_customer_pareto"),
            "nps": fetch_scalar(
                conn, "SELECT SUM(nps * nps_responses) / SUM(nps_responses) "
                      "FROM gold_nps_monthly"),
        }
        cac_sql = text(
            "SELECT (SELECT SUM(spend) FROM gold_channel_cac_monthly) / "
            "(SELECT COUNT(*) FROM silver_customers "
            "WHERE acquisition_channel IN :paid)"
        ).bindparams(bindparam("paid", expanding=True))
        kpis["paid_cac"] = float(
            conn.execute(cac_sql, {"paid": settings.PAID_CHANNELS}).scalar_one())
    return kpis


def compute_rfm(engine) -> pd.DataFrame:
    """Segment summary — same scoring and rule grid as notebook 02."""
    cust = pd.read_sql(
        "SELECT customer_id, orders_completed, revenue_completed, "
        "last_completed_ts FROM gold_customer_summary "
        "WHERE orders_completed > 0", engine, parse_dates=["last_completed_ts"])
    ref_date = cust["last_completed_ts"].max() + pd.Timedelta(days=1)
    rfm = pd.DataFrame({
        "recency_days": (ref_date - cust["last_completed_ts"]).dt.days,
        "frequency": cust["orders_completed"].astype(int),
        "monetary": cust["revenue_completed"].astype(float),
    })
    rfm["R"] = pd.qcut(rfm["recency_days"], 5, labels=[5, 4, 3, 2, 1]).astype(int)
    rfm["F"] = rfm["frequency"].clip(upper=5).astype(int)
    rules = [
        ("Champions", (rfm.R >= 4) & (rfm.F >= 4)),
        ("Loyal", (rfm.R >= 3) & (rfm.F >= 3)),
        ("Can't Lose", (rfm.R <= 2) & (rfm.F >= 4)),
        ("At Risk", (rfm.R <= 2) & (rfm.F >= 2)),
        ("Potential Loyalist", (rfm.R >= 4) & (rfm.F == 2)),
        ("New Customers", (rfm.R >= 4) & (rfm.F == 1)),
        ("Promising", (rfm.R == 3) & (rfm.F == 1)),
        ("Need Attention", (rfm.R == 3) & (rfm.F == 2)),
        ("Hibernating", (rfm.R <= 2) & (rfm.F == 1)),
    ]
    seg = pd.Series(pd.NA, index=rfm.index, dtype="object")
    for name, mask in rules:
        seg = seg.mask(seg.isna() & mask, name)
    if seg.isna().any():
        raise RuntimeError("RFM rule grid left customers unmapped")
    rfm["segment"] = seg
    summary = (rfm.groupby("segment")
               .agg(customers=("segment", "size"),
                    revenue=("monetary", "sum"),
                    avg_revenue=("monetary", "mean"),
                    avg_recency_days=("recency_days", "mean"),
                    avg_frequency=("frequency", "mean"))
               .sort_values("revenue", ascending=False)
               .reset_index())
    summary["customer_share"] = summary["customers"] / summary["customers"].sum()
    summary["revenue_share"] = summary["revenue"] / summary["revenue"].sum()
    return summary


def compute_ab(engine) -> dict:
    """free_shipping_threshold stats — same formulas as notebook 04."""
    ab = pd.read_sql("SELECT variant, converted_flag "
                     "FROM silver_ab_test_assignments", engine)
    n = ab.groupby("variant").size()
    conv = ab.groupby("variant")["converted_flag"].mean()
    n_c, n_t = int(n["control"]), int(n["treatment"])
    p_c, p_t = float(conv["control"]), float(conv["treatment"])
    _, srm_p = stats.chisquare([n_c, n_t])
    p_pool = (n_c * p_c + n_t * p_t) / (n_c + n_t)
    se_pool = np.sqrt(p_pool * (1 - p_pool) * (1 / n_c + 1 / n_t))
    z = (p_t - p_c) / se_pool
    p_value = 2 * stats.norm.sf(abs(z))
    lift = p_t / p_c - 1
    se_log_rr = np.sqrt((1 - p_c) / (n_c * p_c) + (1 - p_t) / (n_t * p_t))
    ci = (np.exp(np.log(p_t / p_c) - 1.96 * se_log_rr) - 1,
          np.exp(np.log(p_t / p_c) + 1.96 * se_log_rr) - 1)
    return {"n_control": n_c, "n_treatment": n_t, "conv_control": p_c,
            "conv_treatment": p_t, "srm_p": float(srm_p), "z": float(z),
            "p_value": float(p_value), "lift": float(lift),
            "ci_low": float(ci[0]), "ci_high": float(ci[1])}


def as_float(df: pd.DataFrame) -> pd.DataFrame:
    """MySQL DECIMAL comes back as object — coerce numerics for openpyxl."""
    out = df.copy()
    for col in out.columns:
        if out[col].dtype == object:
            coerced = pd.to_numeric(out[col], errors="coerce")
            if coerced.notna().sum() == out[col].notna().sum():
                out[col] = coerced
    return out


def write_kv_sheet(writer, sheet: str, title: str, rows: list) -> None:
    """Label/value/format rows starting at A3; labels are the validation keys."""
    pd.DataFrame().to_excel(writer, sheet_name=sheet)  # create empty sheet
    ws = writer.book[sheet]
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    for i, (label, value, fmt) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=label).font = HEADER_FONT
        cell = ws.cell(row=i, column=2, value=value)
        cell.number_format = fmt
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 16


def add_chart(ws, chart, data_ref, cats_ref, anchor: str, title: str) -> None:
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.title = title
    chart.height, chart.width = 8, 18
    ws.add_chart(chart, anchor)


def build(engine) -> dict:
    kpis = compute_kpis(engine)
    rfm = compute_rfm(engine)
    ab = compute_ab(engine)
    monthly = as_float(pd.read_sql(
        "SELECT * FROM gold_revenue_trend ORDER BY order_month", engine))
    funnel = as_float(pd.read_sql(
        "SELECT * FROM gold_funnel ORDER BY step_order", engine))
    cac = as_float(pd.read_sql(
        "SELECT channel, SUM(spend) AS spend, SUM(clicks) AS clicks, "
        "SUM(attributed_signups) AS signups, "
        "SUM(spend) / SUM(attributed_signups) AS cac "
        "FROM gold_channel_cac_monthly GROUP BY channel ORDER BY cac DESC",
        engine))
    cohort = as_float(pd.read_sql(
        "SELECT cohort_month, months_since_first, retention_rate "
        "FROM gold_cohort_retention", engine))
    cohort_mx = cohort.pivot(index="cohort_month",
                             columns="months_since_first",
                             values="retention_rate").sort_index()
    top = as_float(pd.read_sql(
        "SELECT category, category_rank, product_name, revenue, units, "
        "category_share FROM gold_top_products WHERE category_rank <= 10 "
        "ORDER BY category, category_rank", engine))
    nps = as_float(pd.read_sql(
        "SELECT * FROM gold_nps_monthly ORDER BY review_month", engine))
    data_window = f"{monthly['order_month'].iloc[0]} – {monthly['order_month'].iloc[-1]}"

    with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as writer:
        write_kv_sheet(writer, "Executive Summary",
                       f"ShopSphere — Executive Summary ({data_window})", [
            ("Total completed revenue", kpis["total_revenue"], "#,##0.00"),
            ("Completed orders", kpis["completed_orders"], "#,##0"),
            ("Average order value (AOV)", kpis["aov"], "#,##0.00"),
            ("Unique buyers", kpis["buyers"], "#,##0"),
            ("One-time buyer rate", kpis["one_time_rate"], "0.0%"),
            ("Cart abandonment", kpis["abandonment"], "0.0%"),
            ("Top-20% buyers' revenue share", kpis["top20_share"], "0.0%"),
            ("Blended paid CAC", kpis["paid_cac"], "#,##0.00"),
            ("NPS (response-weighted)", kpis["nps"], "0.0"),
            ("A/B free-shipping relative lift", ab["lift"], "+0.0%"),
            ("A/B p-value", ab["p_value"], "0.00E+00"),
        ])

        monthly.to_excel(writer, sheet_name="Monthly Trend", index=False)
        ws = writer.book["Monthly Trend"]
        n_rows = len(monthly) + 1
        add_chart(ws, LineChart(),
                  Reference(ws, min_col=2, min_row=1, max_row=n_rows),
                  Reference(ws, min_col=1, min_row=2, max_row=n_rows),
                  f"A{n_rows + 2}", "Monthly completed revenue")

        funnel.to_excel(writer, sheet_name="Funnel", index=False)
        ws = writer.book["Funnel"]
        n_rows = len(funnel) + 1
        add_chart(ws, BarChart(),
                  Reference(ws, min_col=3, min_row=1, max_row=n_rows),
                  Reference(ws, min_col=2, min_row=2, max_row=n_rows),
                  f"A{n_rows + 2}", "Funnel events by step")

        cac.to_excel(writer, sheet_name="Channel CAC", index=False)
        ws = writer.book["Channel CAC"]
        n_rows = len(cac) + 1
        add_chart(ws, BarChart(),
                  Reference(ws, min_col=5, min_row=1, max_row=n_rows),
                  Reference(ws, min_col=1, min_row=2, max_row=n_rows),
                  f"A{n_rows + 2}", "CAC by channel (spend / attributed signups)")

        cohort_mx.to_excel(writer, sheet_name="Cohort Retention")
        ws = writer.book["Cohort Retention"]
        last_col = get_column_letter(cohort_mx.shape[1] + 1)
        rng = f"B2:{last_col}{cohort_mx.shape[0] + 1}"
        for row in ws[rng]:
            for cell in row:
                cell.number_format = "0%"
        ws.conditional_formatting.add(rng, ColorScaleRule(
            start_type="num", start_value=0, start_color="FFFFFF",
            end_type="num", end_value=0.15, end_color="1F77B4"))

        top.to_excel(writer, sheet_name="Top Products", index=False)
        rfm.to_excel(writer, sheet_name="RFM Segments", index=False)
        nps.to_excel(writer, sheet_name="NPS", index=False)

        write_kv_sheet(writer, "AB Test",
                       "free_shipping_threshold — fixed-horizon analysis", [
            ("Assignments (control)", ab["n_control"], "#,##0"),
            ("Assignments (treatment)", ab["n_treatment"], "#,##0"),
            ("SRM chi-square p", ab["srm_p"], "0.0000"),
            ("Conversion (control)", ab["conv_control"], "0.00%"),
            ("Conversion (treatment)", ab["conv_treatment"], "0.00%"),
            ("Relative lift", ab["lift"], "+0.0%"),
            ("Lift 95% CI low", ab["ci_low"], "+0.0%"),
            ("Lift 95% CI high", ab["ci_high"], "+0.0%"),
            ("z statistic", ab["z"], "0.00"),
            ("p-value", ab["p_value"], "0.00E+00"),
        ])

        for sheet in writer.book.sheetnames:
            for col_cells in writer.book[sheet].columns:
                letter = col_cells[0].column_letter
                dim = writer.book[sheet].column_dimensions[letter]
                if dim.width is None or dim.width < 14:
                    dim.width = 14
    return {"kpis": kpis, "ab": ab, "rfm_customers": int(rfm["customers"].sum())}


def read_kv(ws) -> dict:
    return {ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value
            for r in range(3, ws.max_row + 1)
            if ws.cell(row=r, column=1).value}


def validate(engine, built: dict) -> bool:
    """Re-open the workbook; assert headline cells against fresh SQL."""
    wb = load_workbook(OUT_PATH, data_only=True)
    summary = read_kv(wb["Executive Summary"])
    ab_sheet = read_kv(wb["AB Test"])
    rfm_ws = wb["RFM Segments"]
    header = [c.value for c in rfm_ws[1]]
    cust_col = header.index("customers") + 1
    rfm_total = sum(rfm_ws.cell(row=r, column=cust_col).value
                    for r in range(2, rfm_ws.max_row + 1))

    with engine.connect() as conn:
        fresh_rev = fetch_scalar(
            conn, "SELECT SUM(order_revenue) FROM gold_order_revenue "
                  "WHERE order_status = 'completed'")
        fresh_buyers = fetch_scalar(conn,
                                    "SELECT COUNT(*) FROM gold_customer_pareto")

    checks = [
        ("Total revenue == SQL",
         abs(summary["Total completed revenue"] - fresh_rev) < 0.01,
         f"{summary['Total completed revenue']:,.2f} vs {fresh_rev:,.2f}"),
        ("AOV within 0.5% of gate value",
         abs(summary["Average order value (AOV)"] - AOV_EXPECTED)
         <= AOV_EXPECTED * REL_TOLERANCE,
         f"{summary['Average order value (AOV)']:.4f} vs {AOV_EXPECTED}"),
        ("RFM buyers == gold_customer_pareto",
         rfm_total == int(fresh_buyers),
         f"{rfm_total} vs {int(fresh_buyers)}"),
        ("A/B lift within 0.005 of calibration",
         abs(ab_sheet["Relative lift"] - AB_LIFT_EXPECTED) < 0.005,
         f"{ab_sheet['Relative lift']:.4f} vs {AB_LIFT_EXPECTED}"),
        ("Sheets present",
         set(wb.sheetnames) >= {"Executive Summary", "Monthly Trend", "Funnel",
                                "Channel CAC", "Cohort Retention",
                                "Top Products", "RFM Segments", "NPS",
                                "AB Test"},
         ", ".join(wb.sheetnames)),
    ]
    ok = True
    for name, passed, detail in checks:
        ok = ok and passed
        print(f"  {'PASS' if passed else 'FAIL'} {name}: {detail}")
    return ok


def main() -> int:
    engine = get_engine()
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    built = build(engine)
    print(f"Wrote {OUT_PATH.name} "
          f"(revenue {built['kpis']['total_revenue']:,.2f}, "
          f"AOV {built['kpis']['aov']:.4f}, "
          f"buyers {built['kpis']['buyers']:,.0f}, "
          f"lift {built['ab']['lift']:+.3%})")
    ok = validate(engine, built)
    print("Workbook gate:", "PASS (5/5)" if ok else "FAIL")
    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main())
